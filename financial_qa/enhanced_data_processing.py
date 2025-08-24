#!/usr/bin/env python3
"""
Enhanced Data Processing Module
Handles multiple document types (PDF, Excel, HTML) with OCR, cleaning, and Q/A generation
"""

import re
import json
import logging
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
import pandas as pd
from PIL import Image
import pytesseract
import pdfplumber
from openpyxl import load_workbook
import tabula
from bs4 import BeautifulSoup
import PyPDF2

from .config import RAW_DIR, PROCESSED_DIR, SMALL_CHUNK_CHARS, LARGE_CHUNK_CHARS, CHUNK_OVERLAP_CHARS, ROOT_DIR

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Document patterns
DOC_PATTERNS = {
    "pdf": "**/*.pdf",
    "excel": "**/*.xlsx",
    "html": "**/*.html",
    "docx": "**/*.docx"
}

# Financial section keywords for logical segmentation
FINANCIAL_SECTIONS = {
    "income_statement": [
        "revenue", "revenues", "sales", "income", "earnings", "profit", "loss",
        "operating income", "net income", "gross profit", "ebitda", "ebit"
    ],
    "balance_sheet": [
        "assets", "liabilities", "equity", "cash", "inventory", "accounts receivable",
        "debt", "capital", "retained earnings", "total assets", "total liabilities"
    ],
    "cash_flow": [
        "cash flow", "operating activities", "investing activities", "financing activities",
        "net cash", "cash and cash equivalents", "free cash flow"
    ],
    "operational": [
        "operations", "operating expenses", "research and development", "r&d",
        "general and administrative", "marketing", "sales expenses"
    ]
}


class DocumentProcessor:
    """Handles multiple document types with OCR and parsing"""
    
    def __init__(self, root_dir: Path = ROOT_DIR):
        self.root_dir = root_dir
        self.processed_texts = []
        self.sectioned_texts = []
        
    def discover_documents(self) -> Dict[str, List[Path]]:
        """Find all supported document types"""
        documents = {}
        
        for doc_type, pattern in DOC_PATTERNS.items():
            files = list(self.root_dir.glob(pattern))
            if files:
                documents[doc_type] = sorted(files)
                logger.info(f"Found {len(files)} {doc_type.upper()} files")
        
        return documents
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract text from PDF with OCR fallback"""
        try:
            # Try pdfplumber first (better for text-based PDFs)
            with pdfplumber.open(pdf_path) as pdf:
                text_parts = []
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text and len(page_text.strip()) > 50:  # Meaningful text
                        text_parts.append(page_text)
                    else:
                        # Try OCR for image-based pages
                        ocr_text = self._ocr_page_image(page)
                        if ocr_text:
                            text_parts.append(ocr_text)
                
                return "\n".join(text_parts)
                
        except Exception as e:
            logger.warning(f"pdfplumber failed for {pdf_path}: {e}")
            return self._extract_pdf_with_pypdf2(pdf_path)
    
    def _extract_pdf_with_pypdf2(self, pdf_path: Path) -> str:
        """Fallback PDF extraction using PyPDF2"""
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text_parts = []
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                return "\n".join(text_parts)
        except Exception as e:
            logger.error(f"PyPDF2 extraction failed for {pdf_path}: {e}")
            return ""
    
    def _ocr_page_image(self, page) -> str:
        """Extract text from page image using OCR"""
        try:
            # Convert page to image
            img = page.to_image()
            if img:
                # Convert to PIL Image
                pil_img = Image.fromarray(img.original)
                # Extract text using OCR
                text = pytesseract.image_to_string(pil_img)
                return text.strip()
        except Exception as e:
            logger.debug(f"OCR failed: {e}")
        return ""
    
    def extract_text_from_excel(self, excel_path: Path) -> str:
        """Extract text from Excel files"""
        try:
            workbook = load_workbook(excel_path, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = f"\n=== SHEET: {sheet_name} ===\n"
                
                # Extract text from cells
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                
                text_parts.append(sheet_text)
            
            return "\n".join(text_parts)
            
        except Exception as e:
            logger.error(f"Excel extraction failed for {excel_path}: {e}")
            return ""
    
    def extract_text_from_html(self, html_path: Path) -> str:
        """Extract text from HTML files"""
        try:
            with open(html_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
            
            soup = BeautifulSoup(content, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Extract text
            text = soup.get_text()
            
            # Clean up whitespace
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            return text
            
        except Exception as e:
            logger.error(f"HTML extraction failed for {html_path}: {e}")
            return ""
    
    def extract_text_from_docx(self, docx_path: Path) -> str:
        """Extract text from Word documents"""
        try:
            from docx import Document
            doc = Document(docx_path)
            text_parts = []
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_parts.append(paragraph.text)
            
            return "\n".join(text_parts)
            
        except Exception as e:
            logger.error(f"DOCX extraction failed for {docx_path}: {e}")
            return ""
    
    def clean_text(self, text: str) -> str:
        """Clean text by removing noise, headers, footers, page numbers"""
        if not text:
            return ""
        
        # Remove page numbers
        text = re.sub(r'\b(?:Page|page)\s+\d+\b', '', text)
        text = re.sub(r'\b\d+\s+of\s+\d+\b', '', text)
        
        # Remove common headers/footers
        text = re.sub(r'Rail Vision.*?Financial Results.*?\n', '', text, flags=re.IGNORECASE | re.MULTILINE)
        text = re.sub(r'Confidential.*?\n', '', text, flags=re.IGNORECASE | re.MULTILINE)
        
        # Remove excessive whitespace
        text = re.sub(r'\n+', '\n', text)
        text = re.sub(r' +', ' ', text)
        
        # Remove special characters that might cause issues
        text = re.sub(r'[^\w\s\.\,\$\-\%\(\)]', ' ', text)
        
        # Fix common financial formatting issues
        text = re.sub(r'(\d+),(\d{3})', r'\1,\2', text)  # Fix comma formatting
        text = re.sub(r'\$(\d+)', r'$\1', text)  # Fix dollar sign spacing
        
        return text.strip()
    
    def segment_into_sections(self, text: str) -> List[Dict[str, Any]]:
        """Segment text into logical financial sections"""
        sections = []
        
        # Split text into paragraphs
        paragraphs = text.split('\n\n')
        
        current_section = {"type": "general", "content": [], "keywords": []}
        
        for paragraph in paragraphs:
            if not paragraph.strip():
                continue
            
            # Determine section type based on keywords
            section_type = self._classify_section(paragraph)
            
            if section_type != current_section["type"]:
                # Save current section if it has content
                if current_section["content"]:
                    sections.append(current_section.copy())
                
                # Start new section
                current_section = {
                    "type": section_type,
                    "content": [paragraph],
                    "keywords": self._extract_keywords(paragraph)
                }
            else:
                current_section["content"].append(paragraph)
                current_section["keywords"].extend(self._extract_keywords(paragraph))
        
        # Add final section
        if current_section["content"]:
            sections.append(current_section)
        
        return sections
    
    def _classify_section(self, text: str) -> str:
        """Classify text into financial sections"""
        text_lower = text.lower()
        
        for section_name, keywords in FINANCIAL_SECTIONS.items():
            if any(keyword in text_lower for keyword in keywords):
                return section_name
        
        return "general"
    
    def _extract_keywords(self, text: str) -> List[str]:
        """Extract financial keywords from text"""
        text_lower = text.lower()
        keywords = []
        
        for section_name, section_keywords in FINANCIAL_SECTIONS.items():
            for keyword in section_keywords:
                if keyword in text_lower:
                    keywords.append(keyword)
        
        return list(set(keywords))
    
    def generate_qa_pairs(self, processed_texts: List[Dict]) -> List[Dict[str, str]]:
        """Generate question-answer pairs from processed financial data"""
        qa_pairs = []
        
        # Template questions for different financial aspects
        qa_templates = [
            # Revenue questions
            ("What was the total revenue in {year}?", "revenue", "extract_revenue"),
            ("What were the quarterly revenues for {year}?", "revenue", "extract_quarterly_revenue"),
            ("What was the revenue in Q1 {year}?", "revenue", "extract_quarterly_revenue"),
            ("What was the revenue in Q2 {year}?", "revenue", "extract_quarterly_revenue"),
            ("What was the revenue in Q3 {year}?", "revenue", "extract_quarterly_revenue"),
            ("What was the revenue in Q4 {year}?", "revenue", "extract_quarterly_revenue"),
            ("How much revenue did Rail Vision generate in {year}?", "revenue", "extract_revenue"),
            ("What was the annual revenue for {year}?", "revenue", "extract_revenue"),
            
            # Profit/Loss questions
            ("What was the net income in {year}?", "net_income", "extract_net_income"),
            ("What was the operating profit in {year}?", "operating_profit", "extract_operating_profit"),
            ("What were the total expenses in {year}?", "expenses", "extract_expenses"),
            ("What was the net loss in {year}?", "net_loss", "extract_net_loss"),
            ("What was the operating loss in {year}?", "operating_loss", "extract_operating_loss"),
            ("What was the gross profit in {year}?", "gross_profit", "extract_gross_profit"),
            ("What was the EBITDA in {year}?", "ebitda", "extract_ebitda"),
            
            # Balance sheet questions
            ("What were the total assets at the end of {year}?", "assets", "extract_assets"),
            ("What was the cash position at the end of {year}?", "cash", "extract_cash"),
            ("What were the total liabilities in {year}?", "liabilities", "extract_liabilities"),
            ("What was the shareholders' equity in {year}?", "equity", "extract_equity"),
            ("What was the working capital in {year}?", "working_capital", "extract_working_capital"),
            ("What was the debt level in {year}?", "debt", "extract_debt"),
            
            # Operational questions
            ("What were the R&D expenses in {year}?", "r&d", "extract_rnd_expenses"),
            ("What were the administrative expenses in {year}?", "admin", "extract_admin_expenses"),
            ("What were the general expenses in {year}?", "general_expenses", "extract_general_expenses"),
            ("How many employees did the company have in {year}?", "employees", "extract_employee_count"),
            ("What were the marketing expenses in {year}?", "marketing", "extract_marketing_expenses"),
            ("What were the sales expenses in {year}?", "sales_expenses", "extract_sales_expenses"),
            
            # Growth/Performance questions
            ("What was the revenue growth rate in {year}?", "growth", "extract_growth_rate"),
            ("What was the profit margin in {year}?", "margin", "extract_profit_margin"),
            ("What were the key business highlights in {year}?", "highlights", "extract_highlights"),
            ("What was the market expansion in {year}?", "market_expansion", "extract_market_expansion"),
            ("What were the key orders received in {year}?", "orders", "extract_orders"),
            ("What was the customer acquisition in {year}?", "customers", "extract_customers"),
            
            # Cash Flow questions
            ("What was the operating cash flow in {year}?", "operating_cash_flow", "extract_operating_cash_flow"),
            ("What was the investing cash flow in {year}?", "investing_cash_flow", "extract_investing_cash_flow"),
            ("What was the financing cash flow in {year}?", "financing_cash_flow", "extract_financing_cash_flow"),
            ("What was the free cash flow in {year}?", "free_cash_flow", "extract_free_cash_flow"),
            
            # Technology & Innovation questions
            ("What patents were filed in {year}?", "patents", "extract_patents"),
            ("What new products were launched in {year}?", "products", "extract_products"),
            ("What R&D milestones were achieved in {year}?", "rd_milestones", "extract_rd_milestones"),
            ("What technology partnerships were formed in {year}?", "partnerships", "extract_partnerships"),
            
            # Market & Competition questions
            ("What markets did Rail Vision enter in {year}?", "market_entry", "extract_market_entry"),
            ("What were the competitive advantages in {year}?", "competitive_advantages", "extract_competitive_advantages"),
            ("What regulatory approvals were received in {year}?", "regulatory", "extract_regulatory"),
        ]
        
        # Extract years from the data
        years = self._extract_years_from_texts(processed_texts)
        
        # Add more years if we don't have enough
        if len(years) < 5:
            years.extend(['2021', '2025', '2026'])
        
        for template_question, category, extractor in qa_templates:
            for year in years:
                try:
                    question = template_question.format(year=year)
                    
                    # Try to find answer in the data
                    answer = self._extract_answer(question, processed_texts, year, extractor)
                    
                    if answer:
                        qa_pairs.append({
                            "question": question,
                            "answer": answer,
                            "category": category,
                            "year": year
                        })
                except Exception as e:
                    logger.warning(f"Failed to generate Q/A for template '{template_question}' with year '{year}': {e}")
                    continue
        
        # Add some general questions
        general_qa = [
            ("What is Rail Vision's main business?", "Rail Vision is a technology company focused on railway safety and efficiency solutions."),
            ("What are the main revenue sources?", "Revenue comes from demonstration projects, pilot programs, and commercial contracts."),
            ("What markets does Rail Vision operate in?", "Rail Vision operates in international markets including the US, Australia, and Europe."),
            ("What is the company's technology focus?", "Rail Vision focuses on computer vision and AI technology for railway applications."),
            ("What are the main challenges mentioned in reports?", "Challenges include market penetration, regulatory approvals, and competition."),
            ("What is Rail Vision's competitive advantage?", "Rail Vision has proprietary AI-based railway detection technology and industry expertise."),
            ("What is the company's growth strategy?", "Rail Vision focuses on expanding into new markets, developing new products, and forming strategic partnerships."),
            ("What is Rail Vision's market position?", "Rail Vision is an early commercialization stage company in the railway safety technology market."),
            ("What are Rail Vision's key products?", "Key products include MainLine systems, Shunting Yard systems, and the D.A.S.H. SaaS platform."),
            ("What is Rail Vision's customer base?", "Customers include railway operators, mining companies, and maintenance firms globally."),
            ("What is Rail Vision's funding status?", "Rail Vision has secured over $34 million in funding through various financing transactions."),
            ("What is Rail Vision's regulatory status?", "Rail Vision has received regulatory approvals from Israel Railways and other authorities."),
            ("What is Rail Vision's intellectual property?", "Rail Vision holds patents for AI-powered obstacle detection and other railway safety technologies."),
            ("What is Rail Vision's expansion plan?", "Rail Vision plans to expand into new markets including India and Latin America."),
            ("What is Rail Vision's technology roadmap?", "Rail Vision is developing autonomous train capabilities and enhanced safety systems."),
        ]
        
        for question, answer in general_qa:
            qa_pairs.append({
                "question": question,
                "answer": answer,
                "category": "general",
                "year": "N/A"
            })
        
        return qa_pairs
    
    def _extract_years_from_texts(self, processed_texts: List[Dict]) -> List[str]:
        """Extract years mentioned in the financial data"""
        years = set()
        
        for text_data in processed_texts:
            text = text_data.get("text", "")
            # Find years in format 20XX
            year_matches = re.findall(r'20[12][0-9]', text)
            years.update(year_matches)
        
        return sorted(list(years))
    
    def _extract_answer(self, question: str, processed_texts: List[Dict], year: str, extractor: str) -> Optional[str]:
        """Extract answer based on question and extractor method"""
        question_lower = question.lower()
        
        for text_data in processed_texts:
            text = text_data.get("text", "")
            if year in text:
                # Look for relevant information based on question type
                if "revenue" in question_lower or "revenues" in question_lower:
                    # More targeted approach for revenue
                    revenue = self._extract_revenue_for_year(text, year)
                    if revenue:
                        return revenue
                
                elif "net income" in question_lower or "profit" in question_lower:
                    # Look for net income/loss
                    year_index = text.find(year)
                    if year_index != -1:
                        start = max(0, year_index - 500)
                        end = min(len(text), year_index + 500)
                        context = text[start:end]
                        
                        profit_patterns = [
                            r'net\s+loss[^$]*?\$([0-9,]+)',
                            r'net\s+income[^$]*?\$([0-9,]+)',
                            r'GAAP\s+net\s+loss[^$]*?\$([0-9,]+)',
                        ]
                        
                        for pattern in profit_patterns:
                            matches = re.findall(pattern, context, re.IGNORECASE)
                            if matches:
                                return f"${matches[0]}"
                
                elif "assets" in question_lower:
                    # Look for total assets
                    year_index = text.find(year)
                    if year_index != -1:
                        start = max(0, year_index - 500)
                        end = min(len(text), year_index + 500)
                        context = text[start:end]
                        
                        assets_patterns = [
                            r'total\s+assets[^$]*?\$([0-9,]+)',
                            r'assets[^$]*?\$([0-9,]+)',
                        ]
                        
                        for pattern in assets_patterns:
                            matches = re.findall(pattern, context, re.IGNORECASE)
                            if matches:
                                return f"${matches[0]}"
                
                elif "cash" in question_lower:
                    # Look for cash position
                    year_index = text.find(year)
                    if year_index != -1:
                        start = max(0, year_index - 500)
                        end = min(len(text), year_index + 500)
                        context = text[start:end]
                        
                        cash_patterns = [
                            r'cash\s+and\s+cash\s+equivalents[^$]*?\$([0-9,]+)',
                            r'cash[^$]*?\$([0-9,]+)',
                        ]
                        
                        for pattern in cash_patterns:
                            matches = re.findall(pattern, context, re.IGNORECASE)
                            if matches:
                                return f"${matches[0]}"
        
        return None

    def _extract_revenue_for_year(self, text: str, year: str) -> Optional[str]:
        """Extract revenue specifically for a given year with better context validation"""
        # Look for specific revenue statements for the year
        revenue_patterns = [
            # Pattern 1: "Revenues for the year ended December 31, 2024, increased by 815% to $1.3 million"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?to\s+\$?([0-9,\.]+)\s*million',
            # Pattern 2: "Revenues for the year ended December 31, 2024, were $1,300"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?were\s+\$?([0-9,]+)',
            # Pattern 3: "Revenues for the year ended December 31, 2024, amounted to $1,300"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?amounted\s+to\s+\$?([0-9,]+)',
            # Pattern 4: "Revenues for the year ended December 31, 2024, totaled $1,300"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?totaled\s+\$?([0-9,]+)',
            # Pattern 5: "Revenues for the year ended December 31, 2024, of $1,300"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?of\s+\$?([0-9,]+)',
            # Pattern 6: "Revenues for the year ended December 31, 2024, increased to $1.3 million"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?increased\s+to\s+\$?([0-9,\.]+)\s*million',
            # Pattern 7: "Revenues for the year ended December 31, 2024, $1,300"
            rf'revenues?\s+for\s+the\s+year\s+ended\s+December\s+31,\s*{year}[^$]*?\$?([0-9,]+)',
        ]
        
        for pattern in revenue_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                revenue = match.group(1)
                # Handle "million" conversion
                if 'million' in text[match.start():match.start()+100].lower():
                    try:
                        if '.' in revenue:
                            revenue = str(int(float(revenue) * 1000000))
                        else:
                            revenue = str(int(revenue) * 1000000)
                    except:
                        pass
                return f"${revenue}"
        
        # If no specific pattern found, try to find revenue in the year's context
        # Look for the year and then search for revenue nearby
        year_index = text.find(year)
        if year_index != -1:
            # Look in a smaller, more focused context
            start = max(0, year_index - 200)
            end = min(len(text), year_index + 200)
            context = text[start:end]
            
            # Look for revenue statements in this focused context
            context_revenue_patterns = [
                r'revenues?\s+\$?([0-9,\.]+)',
                r'revenue\s+\$?([0-9,\.]+)',
                r'\$([0-9,\.]+)\s*million',
                r'\$([0-9,]+)',
            ]
            
            for pattern in context_revenue_patterns:
                matches = re.findall(pattern, context, re.IGNORECASE)
                if matches:
                    revenue = matches[0]
                    # Handle "million" conversion
                    if 'million' in context.lower():
                        try:
                            if '.' in revenue:
                                revenue = str(int(float(revenue) * 1000000))
                            else:
                                revenue = str(int(revenue) * 1000000)
                        except:
                            pass
                    return f"${revenue}"
        
        return None
    
    def process_all_documents(self) -> Tuple[List[Dict], List[Dict]]:
        """Process all documents and return processed texts and Q/A pairs"""
        documents = self.discover_documents()
        
        if not documents:
            logger.warning("No documents found to process")
            return [], []
        
        # Process each document type
        for doc_type, file_paths in documents.items():
            logger.info(f"Processing {len(file_paths)} {doc_type.upper()} files...")
            
            for file_path in file_paths:
                logger.info(f"Processing: {file_path.name}")
                
                # Extract text based on file type
                if doc_type == "pdf":
                    text = self.extract_text_from_pdf(file_path)
                elif doc_type == "excel":
                    text = self.extract_text_from_excel(file_path)
                elif doc_type == "html":
                    text = self.extract_text_from_html(file_path)
                elif doc_type == "docx":
                    text = self.extract_text_from_docx(file_path)
                else:
                    continue
                
                if text:
                    # Clean text
                    cleaned_text = self.clean_text(text)
                    
                    # Segment into sections
                    sections = self.segment_into_sections(cleaned_text)
                    
                    # Store processed data
                    self.processed_texts.append({
                        "source": file_path.name,
                        "type": doc_type,
                        "text": cleaned_text,
                        "sections": sections,
                        "raw_length": len(text),
                        "cleaned_length": len(cleaned_text)
                    })
        
        # Generate Q/A pairs
        qa_pairs = self.generate_qa_pairs(self.processed_texts)
        
        logger.info(f"Processed {len(self.processed_texts)} documents")
        logger.info(f"Generated {len(qa_pairs)} Q/A pairs")
        
        return self.processed_texts, qa_pairs
    
    def save_processed_data(self, output_dir: Path = PROCESSED_DIR):
        """Save processed data and Q/A pairs"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save processed texts
        processed_file = output_dir / "processed_financial_data.json"
        with open(processed_file, 'w', encoding='utf-8') as f:
            json.dump(self.processed_texts, f, indent=2, ensure_ascii=False)
        
        # Save Q/A pairs
        qa_file = output_dir / "qa_pairs.jsonl"
        with open(qa_file, 'w', encoding='utf-8') as f:
            for qa in self.generate_qa_pairs(self.processed_texts):
                f.write(json.dumps(qa, ensure_ascii=False) + '\n')
        
        # Save merged text for chunking
        merged_text = "\n\n".join([data["text"] for data in self.processed_texts])
        merged_file = output_dir / "merged_financial_reports.txt"
        merged_file.write_text(merged_text, encoding='utf-8')
        
        logger.info(f"Saved processed data to {output_dir}")
        logger.info(f"Generated {len(self.generate_qa_pairs(self.processed_texts))} Q/A pairs")


def main():
    """Main processing function"""
    processor = DocumentProcessor()
    
    # Process all documents
    processed_texts, qa_pairs = processor.process_all_documents()
    
    # Save results
    processor.save_processed_data()
    
    print(f"\n✅ Processing complete!")
    print(f"📄 Processed {len(processed_texts)} documents")
    print(f"❓ Generated {len(qa_pairs)} Q/A pairs")
    print(f"💾 Data saved to {PROCESSED_DIR}")


if __name__ == "__main__":
    main()
